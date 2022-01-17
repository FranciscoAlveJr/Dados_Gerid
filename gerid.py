from bs4 import BeautifulSoup as bs
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from time import sleep
import pandas as pd
import glob
import logging
import requests as rq
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet._write_only import WriteOnlyCell
import json
from pypasser import reCaptchaV3
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

green = '0000FF00'
yellow = '00FFFF00'
blue = '0000FFFF'

url2 = 'https://requerimento.inss.gov.br/'
url = 'https://requerimento.inss.gov.br/saginternet/pages/agendamento/selecionarServico.xhtml'
login = ''
senha = ''
serviço = 'ATUALIZAÇÃO DE DADOS CADASTRAIS (ATENDIMENTO A DISTÂNCIA)'

def consulta_endereco(dado):
    session = rq.Session()

    url = f'https://buscacepinter.correios.com.br/app/endereco/carrega-cep-endereco.php'

    headers = {
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept-Language': 'pt-BR,pt;q=0.9,en-US;q=0.8,en;q=0.7',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/95.0.4638.69 Safari/537.36',
        'Host': 'buscacepinter.correios.com.br',
        'Origin': 'https://buscacepinter.correios.com.br',
        'Referer': 'https://buscacepinter.correios.com.br/app/endereco/index.php'
        }

    data = {
        'pagina': '/app/endereco/index.php',
        'cepaux': '',
        'mensagem_alerta':'',
        'endereco': dado,
        'tipoCEP': 'ALL'
    }

    endereco = {}

    try:
        res = session.post(url, headers=headers, data=data)
        dados = json.loads(res.text)['dados'][0]

        endereco['estado'] = dados['uf']
        endereco['cidade'] = dados['localidade']
        endereco['logradouro'] = dados['logradouroDNEC']
        endereco['bairro'] = dados['bairro']
        endereco['cep'] = dados['cep']
    except IndexError:
        print(json.loads(res.text)['mensagem'])

    return endereco

def transparencia(cpf):
    # cb = ''
    # with open('cb.txt', 'r+', encoding='UTF-8') as arq:
    #     read = arq.readlines()
    #     for i in range(len(read)):
    #         cb = read[i].strip()

    anchor = ''

            # try:
    token = reCaptchaV3(anchor)
            #     break
            # except:
            #     continue

    url = 'https://www.portaltransparencia.gov.br/pessoa-fisica/busca/resultado'
    header = {
        'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/95.0.4638.69 Safari/537.36',
        }
    params = {
        'termo': cpf,
        'pagina': '1',
        'tamanhoPagina': '10',
        't':'',
        'tokenRecaptcha': token
        }

    res = rq.get(url, headers=header, params=params, verify=False)

    j_res = json.loads(res.text)
    try:
        id_pessoa = j_res['registros'][0]['skPessoa']
    except:
        return None

    link = f'https://www.portaltransparencia.gov.br/busca/pessoa-fisica/{id_pessoa}'

    res2 = rq.get(link, headers=header, verify=False)
    soup2 = bs(res2.text, 'lxml')

    muni_uf = {}
    localidade = soup2.find('section', {'class':'dados-tabelados'}).find_all('span')[2].text.strip()
    if localidade != '':
        local_list = localidade.split('-')
        muni_uf['cidade'] = local_list[0].strip()
        muni_uf['uf'] = local_list[1].strip()
    
    return muni_uf


def main():
    driver = webdriver.Chrome('chromedriver')
    driver.get(url)

    log = driver.find_element(By.ID, 'username')
    log.send_keys(login)

    wa = WebDriverWait(driver, 3600)
    wa.until(EC.presence_of_element_located((By.ID, 'formAgendarConsultar:servicoDrop_input')))

    while True:
        try:
            wa = WebDriverWait(driver, 2)
            wa.until(EC.presence_of_element_located((By.ID, 'formAgendarConsultar:servicoDrop_input')))

            serv = driver.find_element(By.ID, 'formAgendarConsultar:servicoDrop_input')
            serv.send_keys(serviço)
            sleep(1)
            serv.send_keys(Keys.RETURN)

            sleep(2)
            element = WebDriverWait(driver, 10)
            element.until(EC.presence_of_element_located((By.ID, 'formAgendarConsultar:btnAvancarParaDadosRequerente')))

            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            sleep(1)

            element.until(EC.element_to_be_clickable((By.ID, 'formAgendarConsultar:btnAvancarParaDadosRequerente')))
            av = driver.find_element(By.ID, 'formAgendarConsultar:btnAvancarParaDadosRequerente')
            av.click()
            break
        except Exception as e:
            driver.refresh()
            logging.error(e)

    element.until(EC.presence_of_element_located((By.ID, 'formSugestaoDesistenciaRequerimento:btnAvancarSugestaoDesistenciaRequerimento')))

    while True:
        try:
            element.until(EC.element_to_be_clickable((By.ID, 'formSugestaoDesistenciaRequerimento:btnAvancarSugestaoDesistenciaRequerimento')))
            av2 = driver.find_element(By.ID, 'formSugestaoDesistenciaRequerimento:btnAvancarSugestaoDesistenciaRequerimento')
            av2.click()
            break
        except:
            continue

    r = WebDriverWait(driver, 10)
    r.until(EC.presence_of_element_located((By.ID, 'formAgendarConsultar:cpfInput')))

    cookies = driver.get_cookies()[0]
    sessionid = cookies['name']+'='+cookies['value']

    html = driver.page_source
    soup = bs(html, 'html.parser')

    TOKEN = soup.find('input', {'name':'DTPINFRA_TOKEN'})['value']
    JSF = soup.find('input', {'id':'javax.faces.ViewState'})['value']

    url_drive = driver.current_url
    cid = url_drive[-1]

    driver.quit()

    url_p = f'https://requerimento.inss.gov.br/saginternet/pages/agendamento/registrar/dadosRequerenteEntidadeConveniada.xhtml?cid={cid}'


    header = {
        'Accept': 'application/xml, text/xml, */*; q=0.01',
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept-Language': 'pt-BR,pt;q=0.9,en-US;q=0.8,en;q=0.7',
        'Connection': 'keep-alive',
        'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
        'Cookie': sessionid,
        'Faces-Request': 'partial/ajax',
        'Host': 'requerimento.inss.gov.br',
        'Origin': 'https://requerimento.inss.gov.br',
        'Referer': 'https://requerimento.inss.gov.br/saginternet/pages/agendamento/selecionarServico.xhtml',
        'sec-ch-ua': '"Google Chrome";v="95", "Chromium";v="95", ";Not A Brand";v="99"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-origin',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/95.0.4638.69 Safari/537.36',
        'X-Requested-With': 'XMLHttpRequest'
        }

    chaves = ['NOME', 'CPF', 'DATA DE NASCIMENTO', 'IDADE DER', 'NB', 'ESPÉCIE', 'MOTIVO INDEFERIMENTO', 'LOGRADOURO', 'NÚMERO', 'COMPLEMENTO', 'BAIRRO', 'MUNICÍPIO', 'ESTADO', 'CEP', 'LOGRADOURO2', 'NÚMERO2', 'COMPLEMENTO2', 'BAIRRO2','MUNICÍPIO2','ESTADO2','CEP2', 'CELULAR', 'TELEFONE PRINCIPAL', 'TELEFONE SECUNDÁRIO', 'E-MAIL']

    arquivos_xlsx = glob.glob('read/*')
    last = arquivos_xlsx[-1]

    with open('path/sheet.txt', 'r+', encoding='UTF-8') as t:
        pla = int(t.read())

    for x in range(pla, len(arquivos_xlsx)):
        while True:
            try:
                planilha = pd.read_excel(arquivos_xlsx[x], dtype=str)
                cpf1 = planilha['CPF'].tolist()
                nb1 = planilha['NB'].tolist()
                especie1 = planilha['ESPÉCIE'].tolist()
                motivo_ind1 = planilha['MOTIVO INDEFERIMENTO'].tolist()
                idade1 = planilha['IDADE DER'].tolist()
                logra1 = planilha['LOGRADOURO'].tolist()
                tipo_logra1 = planilha['TIPO LOGRADOURO'].tolist()
                num1 = planilha['NÚMERO'].tolist()
                compl1 = planilha['COMPLEMENTO'].tolist()
                bairro1 = planilha['BAIRRO'].tolist()
                cidade1 = planilha['CIDADE'].tolist()
                uf1 = planilha['UF'].tolist()
                cep1 = planilha['CEP'].tolist()

                arq = arquivos_xlsx[x][5:-5]
                
                with open('path/count.txt', 'r+', encoding='UTF-8') as t:
                    count = t.read()
                if count == '0':
                    print()
                    print(f'Analisando planilha {arq}')
                    print()
                else:
                    print()
                    print(f'Continuando planilha {arq}')
                    print()

                with open('path/sheet.txt', 'w+', encoding='UTF-8') as t:
                    t.write(f'{x}')

                
                with open('path/count.txt', 'r+') as t:
                    ini = int(t.read())

                for i in range(ini, len(cpf1)):
                    while True:
                        try:
                            with open('path/count.txt', 'w+', encoding='UTF-8') as t:
                                t.write(f'{i}')

                            c = cpf1[i]
                            nb = nb1[i]
                            especie = especie1[i]
                            motivo_ind = motivo_ind1[i]
                            idade = idade1[i]

                            orig_logra = logra1[i]
                            orig_tipo_logra = tipo_logra1[i]
                            orig_num = num1[i]
                            orig_compl = compl1[i]
                            orig_bair = bairro1[i]
                            orig_cidade = cidade1[i]
                            orig_uf = uf1[i]
                            orig_cep = cep1[i]

                            if pd.isna(orig_tipo_logra):
                                orig_tipo_logra = None

                            if orig_tipo_logra != None and orig_logra != None:
                                orig_logradouro = f'{orig_tipo_logra} {orig_logra}'
                            elif orig_tipo_logra == None and orig_logra != None:
                                orig_logradouro = orig_logra
                            elif orig_tipo_logra != None and orig_logra == None:
                                orig_logradouro = orig_tipo_logra
                            elif orig_tipo_logra == orig_logra == None:
                                orig_logradouro = 'N/D'

                            
                            c = list(c)

                            while len(c) < 11:
                                if len(c) == 11:
                                    print(c)
                                else:
                                    c = list(c)
                                    c.insert(0, '0')
                            c = ''.join(c)

                            cpf = c[:3]+'.'+c[3:6]+'.'+c[6:9]+'-'+c[-2:]

                            data = {
                                'javax.faces.partial.ajax': 'true',
                                'javax.faces.source': 'formAgendarConsultar:btnConsultarCpfAgendamento',
                                'javax.faces.partial.execute': 'formAgendarConsultar:btnConsultarCpfAgendamento formAgendarConsultar:cpfInput formAgendarConsultar:pnlSalarioMaternidadeWrapper',
                                'javax.faces.partial.render': 'formAgendarConsultar frmBotoes frmFormulariosServico',
                                'formAgendarConsultar:btnConsultarCpfAgendamento': 'formAgendarConsultar:btnConsultarCpfAgendamento',
                                'formAgendarConsultar': 'formAgendarConsultar',
                                'DTPINFRA_TOKEN': TOKEN,
                                'formAgendarConsultar:cpfInput': cpf,
                                'formAgendarConsultar:celularInput': '',
                                'formAgendarConsultar:fixoInput': '',
                                'formAgendarConsultar:telefoneSecundarioInput': '',
                                'formAgendarConsultar:emailInput': '',
                                'formAgendarConsultar:cepPrincipalInput': '',
                                'formAgendarConsultar:inputComboTipoLogradouro':'',
                                'formAgendarConsultar:logradouroPrincipalInput': '',
                                'formAgendarConsultar:numeroPrincipalInput': '',
                                'formAgendarConsultar:complementoPrincipalInput': '',
                                'formAgendarConsultar:bairroPrincipalInput': '',
                                'formAgendarConsultar:ufDrop2': '',
                                'formAgendarConsultar:checkEnderecoSecundario_input': 'on',
                                'formAgendarConsultar:paisEndSecundario': '76',
                                'formAgendarConsultar:cepSecundarioInput': '',
                                'formAgendarConsultar:inputComboTipoLogradouroSecundario': '',
                                'formAgendarConsultar:logradouroSecundarioInput': '',
                                'formAgendarConsultar:numeroSecundarioInput': '',
                                'formAgendarConsultar:complementoSecundarioInput': '',
                                'formAgendarConsultar:bairroSecundarioInput': '',
                                'formAgendarConsultar:ufDropSecundario2': '',
                                'javax.faces.ViewState': JSF
                            }

                            res = rq.post(url_p, headers=header, data=data)

                            soup2 = bs(res.content, 'lxml')

                            nome = soup2.find('input', {'id':'formAgendarConsultar:nomeInput'})
                            nasc = soup2.find('input', {'id':'formAgendarConsultar:nascimentoInput'})

                            # Informações de contato
                            celular = soup2.find('input', {'id':'formAgendarConsultar:celularInput'})
                            tel_prin = soup2.find('input', {'id':'formAgendarConsultar:fixoInput'})
                            tel_sec = soup2.find('input', {'id':'formAgendarConsultar:telefoneSecundarioInput'})
                            email = soup2.find('input', {'id':'formAgendarConsultar:emailInput'})

                            # Endereço principal
                            cep = soup2.find('input', {'id':'formAgendarConsultar:cepPrincipalInput'})
                            tip_log = soup2.find('select', {'id':'formAgendarConsultar:inputComboTipoLogradouro'})
                            logr = soup2.find('input', {'id':'formAgendarConsultar:logradouroPrincipalInput'})
                            numero =soup2.find('input', {'id':'formAgendarConsultar:numeroPrincipalInput'})
                            comp = soup2.find('input', {'id':'formAgendarConsultar:complementoPrincipalInput'})
                            bairro = soup2.find('input', {'id':'formAgendarConsultar:bairroPrincipalInput'})
                            estados = soup2.find('select', {'id':'formAgendarConsultar:ufDrop2'})
                            muni = soup2.find('input', {'id':'formAgendarConsultar:municipioEnderecoContato_input'})

                            # Endereço secundário
                            cep2 = soup2.find('input', {'id':'formAgendarConsultar:cepSecundarioInput'})
                            tip_log2 = soup2.find('select', {'id':'formAgendarConsultar:inputComboTipoLogradouroSecundario'})
                            logr2 = soup2.find('input', {'id':'formAgendarConsultar:logradouroSecundarioInput'})
                            numero2 = soup2.find('input', {'id':'formAgendarConsultar:numeroSecundarioInput'})
                            comp2 = soup2.find('input', {'id':'formAgendarConsultar:complementoSecundarioInput'})
                            bairro2 = soup2.find('input', {'id':'formAgendarConsultar:bairroSecundarioInput'})
                            estados2 = soup2.find('select', {'id':'formAgendarConsultar:ufDropSecundario2'})
                            muni2 = soup2.find('input', {'id':'formAgendarConsultar:municipioEnderecoSecundario_input'})
                            
                            try:
                                tipo_selected = tip_log.find('option', {'selected': 'selected'})
                                if tipo_selected.text == 'Selecione um tipo de logradouro':
                                    tipo_selected = None
                            except:
                                tipo_selected = None

                            try:
                                valor_logr = logr['value']
                            except:
                                valor_logr = None

                            if tipo_selected != None and valor_logr != None:
                                logradouro = f'{tipo_selected.text} {valor_logr}'
                            elif tipo_selected == None and valor_logr != None:
                                logradouro = valor_logr
                            elif tipo_selected != None and valor_logr == None:
                                logradouro = tipo_selected.text
                            elif tipo_selected == valor_logr == None:
                                logradouro = 'N/D'

                            estado_sel = estados.find('option', {'selected': 'selected'})

                            if estado_sel == None:
                                estado = 'N/D'
                            else:
                                estado = estado_sel.text

                            try:
                                tipo_selected2 = tip_log2.find('option', {'selected': 'selected'})
                                if tipo_selected2.text == 'Selecione um tipo de logradouro':
                                    tipo_selected2 = None
                            except:
                                tipo_selected2 = None

                            try:
                                valor_logr2 = logr2['value']
                            except:
                                valor_logr2 = None

                            try:
                                if tipo_selected2 != None and valor_logr2 != None:
                                    logradouro2 = f'{tipo_selected2.text} {valor_logr2}'
                                elif tipo_selected2 == None and valor_logr2 != None:
                                    logradouro2 = valor_logr2
                                elif tipo_selected2 != None and valor_logr2 == None:
                                    logradouro2 = tipo_selected2.text
                                elif tipo_selected2 == valor_logr2 == None:
                                    logradouro2 = 'N/D'
                            except:
                                logradouro2 = 'N/D'

                            try:
                                estado_sel2 = estados2.find('option', {'selected': 'selected'})
                                if estado_sel2 == None:
                                    estado2 = 'N/D'
                                else:
                                    estado2 = estado_sel2.text
                            except:
                                estado2 = 'N/D'

                            tags = [nome, c, nasc, idade, nb, especie, motivo_ind, logradouro, numero, comp, bairro, muni, estado, cep, logradouro2, numero2, comp2, bairro2, muni2, estado2, cep2, celular, tel_prin, tel_sec, email]
                            
                            valores = []

                            for valor in tags:
                                try:
                                    if type(valor) == str:
                                        valores.append(valor)
                                    else:
                                        valores.append(valor['value'])
                                except:
                                    valores.append('N/D')

                            logradouro = valores[7]
                            bairro = valores[10]
                            municipio = valores[11]
                            estado = valores[12]
                            cep = valores[13]
                            endereco1 = [logradouro, bairro, municipio, estado, cep]

                            logradouro2 = valores[14]
                            bairro2 = valores[17]
                            municipio2 = valores[18]
                            estado2 = valores[19]
                            cep2 = valores[20]
                            endereco2 = [logradouro2, bairro2, municipio2, estado2]

                            color = '00FFFFFF'
                            color2 = '00FFFFFF'
                            color3 = '00FFFFFF'

                            if len(cep) == 10:
                                n_cep = ''.join(cep.split('.'))
                                cep = ''.join(n_cep.split('-'))

                            mods = []
                            mods2 = []
                            mods3 = []

                            if 'N/D' in endereco1:
                                if cep != 'N/D':
                                    try:
                                        endereco = consulta_endereco(cep)
                                        color = green
                                        if valores[7] == 'N/D': 
                                            valores[7] = endereco['logradouro']
                                            mods.append(7)
                                        if valores[10] == 'N/D':
                                            valores[10] = endereco['bairro']
                                            mods.append(10)
                                        if valores[11] == 'N/D':
                                            valores[11] = endereco['cidade']
                                            mods.append(11)
                                        if valores[12] == 'N/D':
                                            valores[12] = endereco['estado']
                                            mods.append(12)
                                    except:
                                        pass

                                elif cep == 'N/D' and logradouro != 'N/D' and bairro != 'N/D':
                                    if 'SEM TETO' not in endereco1:
                                        color = green
                                        ende = f'{logradouro} {bairro}'
                                        try:
                                            endereco = consulta_endereco(ende)
                                            if valores[11] == 'N/D':
                                                valores[11] = endereco['cidade']
                                                mods.append(11)
                                            if valores[12] == 'N/D':
                                                valores[12] = endereco['estado']
                                                mods.append(12)
                                            if valores[13] == 'N/D':
                                                valores[13] = endereco['cep']
                                                mods.append(13)
                                        except:
                                            pass

                                elif cep == 'N/D' and logradouro == 'N/D' and bairro == 'N/D':
                                    color = yellow
                                    valores[7] = orig_logradouro
                                    mods.append(7)
                                    valores[8] = orig_num
                                    mods.append(8)
                                    valores[9] = orig_compl
                                    mods.append(9)
                                    valores[10] = orig_bair
                                    mods.append(10)
                                    valores[11] = orig_cidade
                                    mods.append(11)
                                    valores[12] = orig_uf
                                    mods.append(12)
                                    valores[13] = orig_cep
                                    mods.append(13)
                                                    
                            if 'N/D' in endereco2:
                                if len(cep2) == 10:
                                    n_cep = ''.join(cep.split('.'))
                                    cep2 = ''.join(n_cep.split('-'))

                                if cep2 != 'N/D':
                                    try:
                                        color2 = green
                                        endereco = consulta_endereco(cep2)
                                        if valores[14] == 'N/D':
                                            valores[14] = endereco['logradouro']
                                            mods2.append(14)
                                        if valores[17] == 'N/D':
                                            valores[17] = endereco['bairro']
                                            mods2.append(17)
                                        if valores[18] == 'N/D':
                                            valores[18] = endereco['cidade']
                                            mods2.append(18)
                                        if valores[19] == 'N/D':
                                            valores[19] = endereco['estado']
                                            mods2.append(19)
                                    except:
                                        pass

                                elif cep2 == 'N/D' and logradouro2 != 'N/D' and bairro2 != 'N/D':
                                    try:
                                        ende = f'{logradouro2} {bairro2}'
                                        endereco = consulta_endereco(ende)
                                        color2 = green
                                        if valores[18] == 'N/D':
                                            valores[18] = endereco['cidade']
                                            mods2.append(18)
                                        if valores[19] == 'N/D':
                                            valores[19] = endereco['estado']
                                            mods2.append(19)
                                        if valores[20] == 'N/D':
                                            valores[20] = endereco['cep']
                                            mods2.append(20)
                                    except:
                                        pass

                            try:
                                wb = load_workbook(f'results/{arq} RES.xlsx')
                                ws = wb.worksheets[0]
                            except FileNotFoundError:
                                wb = Workbook()
                                ws = wb.active
                                ws.append(chaves)

                            row = []
                            cont = 0
                            if valores[11] == 'N/D' or valores[12] == 'N/D' or valores[11] == valores[12] == 'N/D' or color == yellow:
                                while cont < 5:
                                    muni_uf = transparencia(c)
                                    if muni_uf != None and muni_uf != {}:
                                        break
                                    else:
                                        cont += 1
                                        print('Nada encontrado no portal.\nTentando novamente.')
                                        continue
                                if cont == 5:
                                    color = yellow
                                    if valores[7] == 'N/D':
                                        valores[7] = orig_logradouro
                                        mods.append(7)
                                    if valores[8] == 'N/D':
                                        valores[8] = orig_num
                                        mods.append(8)
                                    if valores[9] == 'N/D':
                                        valores[9] = orig_compl
                                        mods.append(9)
                                    if valores[10] == 'N/D':
                                        valores[10] = orig_bair
                                        mods.append(10)
                                    if valores[11] == 'N/D':
                                        valores[11] = orig_cidade
                                        mods.append(11)
                                    if valores[12] == 'N/D':
                                        valores[12] = orig_uf
                                        mods.append(12)
                                    if valores[13] == 'N/D':
                                        valores[13] = orig_cep
                                        mods.append(13)

                            for f, valor in enumerate(valores):
                                try:
                                    if valor == '':
                                        valor = 'N/D'
                                    if cont < 5:
                                        if f == 11 and valor == 'N/D' or color == yellow and f == 11:
                                            while True:
                                                muni_uf = transparencia(c)
                                                if muni_uf != None and muni_uf != {}:
                                                    color3 = blue
                                                    mods3.append(f)
                                                    valor = muni_uf['cidade']
                                                    break
                                                elif muni_uf == None and muni_uf == {}:
                                                    continue
                                        elif f == 12 and valor == 'N/D' or color == yellow and f == 12:
                                            while True:
                                                muni_uf = transparencia(c)
                                                if  muni_uf != None and muni_uf != {}:
                                                    color3 = blue
                                                    mods3.append(f)
                                                    valor = muni_uf['uf']
                                                    break
                                                elif  muni_uf == None and muni_uf == {}:
                                                    continue

                                    cell = WriteOnlyCell(ws, value=valor)

                                    if f in mods and valor != 'N/D':
                                        cell.fill = PatternFill(fgColor=color, fill_type='solid')

                                    elif f in mods2 and valor != 'N/D':
                                        cell.fill = PatternFill(fgColor=color2, fill_type='solid')
                                    
                                    if f in mods3:
                                        cell.fill = PatternFill(fgColor=color3, fill_type='solid')
                                except IndexError:
                                    continue
                
                                row.append(cell)

                            ws.append(row)
                            wb.save(f'results/{arq} RES.xlsx')
                            
                            print(f'{i+1} - {valores[1]} - {valores[0]}')
                            break
                        except Exception as e:
                            print(e)
                            print(f'CPF: {cpf}')
                            print('Tentando novamente...')
                            sleep(1)
                            continue
                
                with open('path/count.txt', 'w+', encoding='UTF-8') as t:
                    t.write('0')
                print(f'Nova planilha {arq} RES completa!')
                break
            except Exception as e:
                print(e)
                continue

    with open('path/sheet.txt', 'w+', encoding='UTF-8') as t:
        t.write('0')

if __name__=='__main__':
    main()

